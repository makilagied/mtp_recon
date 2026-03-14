"""
DSE MTP vs Bank Statement Reconciliation — Streamlit UI.
Upload MTP trades and bank statement Excel files, map columns, run reconciliation.
"""

import io
import re
import zipfile
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from reconcile import reconcile, ReconciliationResult

# Excel highlighting: green = matched; white = unmatched (bank); yellow = unmatched (MTP control)
FILL_MATCHED = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
FILL_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
FILL_UNMATCHED_MTP = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


def _safe_for_display(df: pd.DataFrame) -> pd.DataFrame:
    """Convert columns that can overflow PyArrow (e.g. large ints) to string for display."""
    if df is None or df.empty:
        return df
    out = df.copy()
    for col in out.columns:
        try:
            dtype = out[col].dtype
            if pd.api.types.is_integer_dtype(dtype):
                out[col] = out[col].astype("object").astype(str)
            elif dtype == object:
                # object columns may contain Python ints too large for C long
                out[col] = out[col].astype(str)
        except (TypeError, ValueError, OverflowError):
            out[col] = out[col].astype(str)
    return out


def _apply_row_highlights(ws, ncols: int, status_col_idx: int):
    """Apply green fill to matched rows, white to unmatched (for bank statement)."""
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=status_col_idx)
        fill = FILL_MATCHED if (cell.value == "Matched") else FILL_WHITE
        for col_idx in range(1, ncols + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill


def _excel_with_highlighted_rows(df: pd.DataFrame, status_column: str = "Recon status") -> bytes:
    """Write dataframe to Excel and fill rows: green if matched, white if unmatched (bank)."""
    buf = io.BytesIO()
    df.to_excel(buf, index=False, engine="openpyxl")
    buf.seek(0)
    wb = load_workbook(buf)
    ws = wb.active
    try:
        status_col_idx = list(df.columns).index(status_column) + 1  # 1-based
    except ValueError:
        status_col_idx = df.shape[1]
    _apply_row_highlights(ws, df.shape[1], status_col_idx)
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


def _excel_with_control_column_highlight(
    df: pd.DataFrame,
    status_column: str = "Recon status",
    control_column: str | None = None,
) -> bytes:
    """Write dataframe to Excel; only the control number column is filled: green if matched, white if unmatched."""
    if control_column is None or control_column not in df.columns:
        control_column = df.columns[0]
    buf = io.BytesIO()
    df.to_excel(buf, index=False, engine="openpyxl")
    buf.seek(0)
    wb = load_workbook(buf)
    ws = wb.active
    try:
        status_col_idx = list(df.columns).index(status_column) + 1
    except ValueError:
        status_col_idx = len(df.columns)
    control_col_idx = list(df.columns).index(control_column) + 1
    for row_idx in range(2, ws.max_row + 1):
        status_cell = ws.cell(row=row_idx, column=status_col_idx)
        fill = FILL_MATCHED if (status_cell.value == "Matched") else FILL_UNMATCHED_MTP
        ws.cell(row=row_idx, column=control_col_idx).fill = fill
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


def _sanitize_sheet_name(name: str, max_len: int = 31) -> str:
    """Excel sheet names: max 31 chars, no : \\ / ? * [ ]."""
    name = re.sub(r'[:\\/?*\[\]]', "_", str(name))
    return name[:max_len] if len(name) > max_len else name or "Sheet"

st.set_page_config(page_title="DSE MTP Reconciliation", layout="wide")
st.title("DSE MTP Trades vs Bank Statement Reconciliation")
st.caption(
    "Match MTP trades to bank statement rows where the DSE control number appears "
    "in the narration/details and the amount paid equals the credit amount."
)

# Session state for uploaded data and result
if "result" not in st.session_state:
    st.session_state.result = None
if "preserved_matched_mtp" not in st.session_state:
    st.session_state.preserved_matched_mtp = None
if "preserved_matched_bank" not in st.session_state:
    st.session_state.preserved_matched_bank = None

# --- File upload ---
st.header("1. Upload files")
col1, col2 = st.columns(2)

with col1:
    mtp_files = st.file_uploader(
        "MTP trades (Excel) — upload one or more",
        type=["xlsx", "xls"],
        key="mtp_upload",
        accept_multiple_files=True,
        help="Each file's first sheet is used. Map control number and amount per file. If a file has a 'Recon status' column (Matched/Unmatched), only Unmatched rows will be reconciled and Matched rows preserved.",
    )
with col2:
    bank_files = st.file_uploader(
        "Bank statement(s) (Excel) — upload one or more",
        type=["xlsx", "xls"],
        key="bank_upload",
        accept_multiple_files=True,
        help="Upload multiple files to reconcile against all statements together. Each file's first sheet is used. If a file has a 'Recon status' column (Matched/Unmatched), only Unmatched rows will be used for matching and Matched rows preserved.",
    )


def load_excel(uploaded_file, sheet_index: int = 0) -> pd.DataFrame | None:
    if uploaded_file is None:
        return None
    try:
        return pd.read_excel(uploaded_file, sheet_name=sheet_index)
    except Exception as e:
        st.error(f"Could not read {getattr(uploaded_file, 'name', 'file')}: {e}")
        return None


def _idx(cols, *prefer):
    for p in prefer:
        if p in cols:
            return cols.index(p)
    return 0


# Store each MTP file as (name, df)
if "mtp_files_data" not in st.session_state:
    st.session_state.mtp_files_data = []

if mtp_files:
    st.session_state.mtp_files_data = []
    for i, f in enumerate(mtp_files):
        name = getattr(f, "name", f"MTP_{i+1}")
        df = load_excel(f)
        if df is not None and not df.empty:
            st.session_state.mtp_files_data.append((name, df))
else:
    st.session_state.mtp_files_data = []
    st.session_state.result = None

# Store each bank statement as (name, df) so we can map columns per file
if "bank_files_data" not in st.session_state:
    st.session_state.bank_files_data = []

if bank_files:
    # (Re)build list of (name, df) for current set of uploads
    st.session_state.bank_files_data = []
    for i, f in enumerate(bank_files):
        name = getattr(f, "name", f"Statement {i+1}")
        df = load_excel(f)
        if df is not None and not df.empty:
            st.session_state.bank_files_data.append((name, df))
else:
    st.session_state.bank_files_data = []
    st.session_state.result = None

mtp_files_data = st.session_state.mtp_files_data
bank_files_data = st.session_state.bank_files_data

# --- Column mapping ---
st.header("2. Map columns")

if mtp_files_data and bank_files_data:
    col_a, col_b = st.columns(2)
    with col_a:
        st.subheader("MTP file(s) — map per file")
        st.caption("Each file can have different column names. Choose control number and amount for each.")
        mtp_control_choices = []
        mtp_amount_choices = []
        for i, (name, df) in enumerate(mtp_files_data):
            mcols = list(df.columns)
            with st.expander(f"**{name}** ({len(df)} rows)", expanded=(i == 0)):
                mi = _idx(mcols, "dse_control_number", "control_number", "reference", "Reference")
                ai = _idx(mcols, "amount_paid", "amount", "credit", "Amount")
                ctrl = st.selectbox(
                    "DSE control number",
                    options=mcols,
                    index=mi,
                    key=f"mtp_ctrl_{i}",
                )
                amt = st.selectbox(
                    "Amount paid",
                    options=mcols,
                    index=ai,
                    key=f"mtp_amt_{i}",
                )
                mtp_control_choices.append(ctrl)
                mtp_amount_choices.append(amt)
                st.dataframe(_safe_for_display(df.head(5)), width="stretch", height=120)
    with col_b:
        st.subheader("Bank statement(s) — map per file")
        st.caption("Each file can have different column names. Choose narration and credit for each.")
        bank_narration_choices = []
        bank_credit_choices = []
        for i, (name, df) in enumerate(bank_files_data):
            bcols = list(df.columns)
            with st.expander(f"**{name}** ({len(df)} rows)", expanded=(i == 0)):
                ni = _idx(bcols, "Narration", "Description", "Details", "Particulars", "Remarks")
                ci = _idx(bcols, "Credit", "credit", "Amount", "credit_amount", "Credit Amount")
                narr = st.selectbox(
                    "Narration / Description / Details",
                    options=bcols,
                    index=ni,
                    key=f"bn_narr_{i}",
                )
                cred = st.selectbox(
                    "Credit amount",
                    options=bcols,
                    index=ci,
                    key=f"bn_cred_{i}",
                )
                bank_narration_choices.append(narr)
                bank_credit_choices.append(cred)
                st.dataframe(_safe_for_display(df.head(5)), width="stretch", height=120)

    # Build combined MTP dataframe with standardized columns
    mtp_combined_frames = []
    for i, (name, df) in enumerate(mtp_files_data):
        part = df.copy()
        part = part.rename(columns={
            mtp_control_choices[i]: "_recon_control",
            mtp_amount_choices[i]: "_recon_amount",
        })
        part["_mtp_source"] = name
        mtp_combined_frames.append(part)
    mtp_df = pd.concat(mtp_combined_frames, ignore_index=True)

    # If MTP has "Recon status" column, only reconcile rows that are not already Matched
    recon_status_col = None
    for c in mtp_df.columns:
        if str(c).strip().lower() == "recon status":
            recon_status_col = c
            break
    preserved_matched_mtp = None
    if recon_status_col is not None:
        # Normalize for comparison (Matched / Unmatched / Multi-match etc.)
        status_vals = mtp_df[recon_status_col].astype(str).str.strip().str.lower()
        is_matched = status_vals == "matched"
        preserved_matched_mtp = mtp_df.loc[is_matched].copy()
        mtp_df = mtp_df.loc[~is_matched].copy()
        if not preserved_matched_mtp.empty:
            st.session_state.preserved_matched_mtp = preserved_matched_mtp
        else:
            st.session_state.preserved_matched_mtp = None
    else:
        st.session_state.preserved_matched_mtp = None

    # Build combined bank dataframe with standardized narration/credit columns
    combined_frames = []
    for i, (name, df) in enumerate(bank_files_data):
        part = df.copy()
        narr_col = bank_narration_choices[i]
        cred_col = bank_credit_choices[i]
        part = part.rename(columns={
            narr_col: "_recon_narration",
            cred_col: "_recon_credit",
        })
        part["_bank_source"] = name
        combined_frames.append(part)
    bank_df = pd.concat(combined_frames, ignore_index=True)

    # If bank has "Recon status" column, only use rows that are not already Matched for reconciliation
    bank_recon_status_col = None
    for c in bank_df.columns:
        if str(c).strip().lower() == "recon status":
            bank_recon_status_col = c
            break
    if bank_recon_status_col is not None:
        status_vals_bank = bank_df[bank_recon_status_col].astype(str).str.strip().str.lower()
        is_matched_bank = status_vals_bank == "matched"
        st.session_state.preserved_matched_bank = bank_df.loc[is_matched_bank].copy()
        bank_df = bank_df.loc[~is_matched_bank].copy()
        if st.session_state.preserved_matched_bank.empty:
            st.session_state.preserved_matched_bank = None
    else:
        st.session_state.preserved_matched_bank = None

    st.header("3. Run reconciliation")
    has_preserved_mtp = st.session_state.preserved_matched_mtp is not None and not st.session_state.preserved_matched_mtp.empty
    has_preserved_bank = st.session_state.preserved_matched_bank is not None and not st.session_state.preserved_matched_bank.empty
    if has_preserved_mtp or has_preserved_bank:
        st.info("Using **Recon status**: only **Unmatched** rows will be reconciled; **Matched** rows from the upload(s) are preserved.")
    if st.button("Reconcile", type="primary"):
        try:
            res = reconcile(
                mtp_df=mtp_df,
                mtp_control_col="_recon_control",
                mtp_amount_col="_recon_amount",
                bank_df=bank_df,
                bank_narration_col="_recon_narration",
                bank_credit_col="_recon_credit",
            )
            st.session_state.result = res
            st.success("Reconciliation completed.")
        except KeyError as e:
            st.error(str(e))
        except Exception as e:
            st.error(f"Reconciliation failed: {e}")

    st.markdown("[☕ Buy me a coffee](https://snippe.me/pay/makilagied)")

    # --- Results ---
    res: ReconciliationResult | None = st.session_state.result
    preserved_matched_mtp = st.session_state.preserved_matched_mtp
    preserved_matched_bank = st.session_state.preserved_matched_bank
    if res is not None:
        st.header("4. Results")

        n_preserved_mtp = len(preserved_matched_mtp) if preserved_matched_mtp is not None and not preserved_matched_mtp.empty else 0
        n_preserved_bank = len(preserved_matched_bank) if preserved_matched_bank is not None and not preserved_matched_bank.empty else 0
        n_preserved = n_preserved_mtp + n_preserved_bank
        n_new_matched = len(res.matched)
        n_matched = n_new_matched + n_preserved
        n_unmatched_mtp = len(res.unmatched_mtp)
        n_unmatched_bank = len(res.unmatched_bank)
        n_multi = len(res.multi_matches)
        total_mtp_rows = sum(len(mdf) for _, mdf in mtp_files_data)
        total_bank_rows = sum(len(bdf) for _, bdf in bank_files_data)
        match_rate_mtp = (n_matched / total_mtp_rows * 100) if total_mtp_rows else 0
        match_rate_bank = (n_matched / total_bank_rows * 100) if total_bank_rows else 0

        # Status callout
        if n_unmatched_mtp == 0 and n_unmatched_bank == 0 and n_multi == 0:
            st.success("All rows reconciled — no unmatched or multi-matches to review.")
        elif n_multi > 0:
            st.warning(f"{n_unmatched_mtp} unmatched MTP, {n_unmatched_bank} unmatched bank, {n_multi} multi-matches — review required.")
        else:
            st.info(f"{n_unmatched_mtp} unmatched MTP and {n_unmatched_bank} unmatched bank rows to review.")

        # Summary metrics
        st.subheader("Summary")
        r1_1, r1_2, r1_3, r1_4 = st.columns(4)
        r1_1.metric("Preserved (MTP)", n_preserved_mtp, help="Rows already Matched from Recon status in uploaded MTP file(s)")
        r1_2.metric("Preserved (Bank)", n_preserved_bank, help="Rows already Matched from Recon status in uploaded bank file(s)")
        r1_3.metric("New matches", n_new_matched, help="Pairs matched in this run")
        r1_4.metric("Total matched", n_matched, help="Preserved + New matches")
        r2_1, r2_2, r2_3 = st.columns(3)
        r2_1.metric("Unmatched MTP", n_unmatched_mtp, help="MTP rows with no matching bank entry")
        r2_2.metric("Unmatched bank", n_unmatched_bank, help="Bank rows not matched to any MTP")
        r2_3.metric("Multi-matches", n_multi, help="MTP rows that matched more than one bank row — review manually")

        # Match rate and volume
        st.subheader("Match rate & volume")
        rate_1, rate_2, rate_3, rate_4 = st.columns(4)
        rate_1.metric("Total MTP rows", total_mtp_rows, help="Rows across all uploaded MTP files")
        rate_2.metric("Total bank rows", total_bank_rows, help="Rows across all uploaded bank files")
        rate_3.metric("MTP match rate", f"{match_rate_mtp:.1f}%", help="Total matched / Total MTP rows")
        rate_4.metric("Bank match rate", f"{match_rate_bank:.1f}%", help="Total matched / Total bank rows")
        st.progress(min(1.0, n_matched / total_mtp_rows) if total_mtp_rows else 0, text="MTP rows matched" if total_mtp_rows else None)

        # Outcome distribution chart
        st.subheader("Outcome distribution")
        outcome_df = pd.DataFrame({
            "Outcome": ["Matched", "Unmatched MTP", "Unmatched bank", "Multi-match"],
            "Count": [n_matched, n_unmatched_mtp, n_unmatched_bank, n_multi],
        }).set_index("Outcome")
        st.bar_chart(outcome_df, height=280)

        # Match breakdown by file (compute rows first for charts)
        mtp_file_names = [name for name, _ in mtp_files_data]
        rows_mtp = []
        for name in mtp_file_names:
            preserved = 0
            if preserved_matched_mtp is not None and "_mtp_source" in preserved_matched_mtp.columns:
                preserved = int((preserved_matched_mtp["_mtp_source"] == name).sum())
            new = 0
            if not res.matched.empty and "_mtp_source" in res.matched.columns:
                new = int((res.matched["_mtp_source"] == name).sum())
            total = preserved + new
            rows_mtp.append({"File": name, "Preserved": preserved, "New": new, "Total": total})
        bank_file_names = [name for name, _ in bank_files_data]
        bank_source_col = "bank__bank_source" if not res.matched.empty and "bank__bank_source" in res.matched.columns else None
        rows_bank = []
        for name in bank_file_names:
            preserved = 0
            if preserved_matched_bank is not None and "_bank_source" in preserved_matched_bank.columns:
                preserved = int((preserved_matched_bank["_bank_source"] == name).sum())
            new = 0
            if bank_source_col:
                new = int((res.matched[bank_source_col] == name).sum())
            total = preserved + new
            rows_bank.append({"File": name, "Preserved": preserved, "New": new, "Total": total})

        st.subheader("Match breakdown by file")
        breakdown_col_mtp, breakdown_col_bank = st.columns(2)
        with breakdown_col_mtp:
            st.caption("**MTP files**")
            if rows_mtp:
                btm_df = pd.DataFrame(rows_mtp)
                st.bar_chart(btm_df.set_index("File")[["Preserved", "New", "Total"]], height=220)
                st.dataframe(btm_df, use_container_width=True, hide_index=True)
            else:
                st.caption("No MTP files.")
        with breakdown_col_bank:
            st.caption("**Bank files**")
            if rows_bank:
                btb_df = pd.DataFrame(rows_bank)
                st.bar_chart(btb_df.set_index("File")[["Preserved", "New", "Total"]], height=220)
                st.dataframe(btb_df, use_container_width=True, hide_index=True)
            else:
                st.caption("No bank files.")

        st.markdown("---")
        with st.expander("Detailed data tables (Matched, Unmatched, Multi-matches)", expanded=True):
            tab1, tab2, tab3, tab4 = st.tabs(
                ["Matched", "Unmatched MTP", "Unmatched bank", "Multi-matches"]
            )
            with tab1:
                if n_preserved_mtp > 0:
                    st.caption("Preserved as Matched — MTP (from Recon status in upload)")
                    st.dataframe(_safe_for_display(preserved_matched_mtp), width="stretch")
                if n_preserved_bank > 0:
                    st.caption("Preserved as Matched — Bank (from Recon status in upload)")
                    st.dataframe(_safe_for_display(preserved_matched_bank), width="stretch")
                if n_preserved > 0:
                    st.markdown("---")
                    st.caption("Newly matched (this run)")
                if not res.matched.empty:
                    st.dataframe(_safe_for_display(res.matched), width="stretch")
                elif n_preserved == 0:
                    st.info("No matched rows.")
            with tab2:
                if not res.unmatched_mtp.empty:
                    st.dataframe(_safe_for_display(res.unmatched_mtp), width="stretch")
                else:
                    st.info("All MTP rows were matched.")
            with tab3:
                if not res.unmatched_bank.empty:
                    st.dataframe(_safe_for_display(res.unmatched_bank), width="stretch")
                else:
                    st.info("All bank rows were matched.")
            with tab4:
                if not res.multi_matches.empty:
                    st.warning(
                        "These MTP rows matched more than one bank row; resolve manually."
                    )
                    st.dataframe(_safe_for_display(res.multi_matches), width="stretch")
                else:
                    st.info("No multi-matches.")

        # Export (include summary and breakdown sheets)
        st.subheader("Export")
        results_buf = io.BytesIO()
        with pd.ExcelWriter(results_buf, engine="openpyxl") as writer:
            # Summary counts and rates
            summary_df = pd.DataFrame([
                {"Metric": "Preserved (MTP)", "Count": n_preserved_mtp},
                {"Metric": "Preserved (Bank)", "Count": n_preserved_bank},
                {"Metric": "New matches", "Count": n_new_matched},
                {"Metric": "Total matched", "Count": n_matched},
                {"Metric": "Unmatched MTP", "Count": n_unmatched_mtp},
                {"Metric": "Unmatched bank", "Count": n_unmatched_bank},
                {"Metric": "Multi-matches", "Count": n_multi},
                {"Metric": "Total MTP rows", "Count": total_mtp_rows},
                {"Metric": "Total bank rows", "Count": total_bank_rows},
                {"Metric": "MTP match rate %", "Count": round(match_rate_mtp, 2)},
                {"Metric": "Bank match rate %", "Count": round(match_rate_bank, 2)},
            ])
            summary_df.to_excel(writer, sheet_name="Summary", index=False)
            # Breakdown by file
            breakdown_mtp_df = pd.DataFrame(rows_mtp) if rows_mtp else pd.DataFrame(columns=["File", "Preserved", "New", "Total"])
            breakdown_mtp_df.to_excel(writer, sheet_name="Breakdown_MTP_files", index=False)
            breakdown_bank_df = pd.DataFrame(rows_bank) if rows_bank else pd.DataFrame(columns=["File", "Preserved", "New", "Total"])
            breakdown_bank_df.to_excel(writer, sheet_name="Breakdown_Bank_files", index=False)
            if n_preserved_mtp > 0:
                preserved_matched_mtp.to_excel(writer, sheet_name="Preserved_Matched_MTP", index=False)
            if n_preserved_bank > 0:
                preserved_matched_bank.to_excel(writer, sheet_name="Preserved_Matched_Bank", index=False)
            res.matched.to_excel(writer, sheet_name="Matched", index=False)
            res.unmatched_mtp.to_excel(writer, sheet_name="Unmatched_MTP", index=False)
            res.unmatched_bank.to_excel(
                writer, sheet_name="Unmatched_Bank", index=False
            )
            res.multi_matches.to_excel(
                writer, sheet_name="Multi_matches", index=False
            )
        results_excel_bytes = results_buf.getvalue()
        st.download_button(
            "Download results (Excel)",
            data=results_excel_bytes,
            file_name="mtp_reconciliation_results.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        # Download each uploaded file individually with reconciliation results (Recon status + highlighting)
        st.subheader("Download uploaded files with results")
        st.caption(
            "Each file you uploaded is available below with a **Recon status** column and row highlighting (green = Matched, yellow/white = Unmatched). "
            "Download files one by one or use the ZIP to get all at once."
        )
        matched_mtp_indices = set(res.matched["_mtp_index"].values) if not res.matched.empty else set()
        if preserved_matched_mtp is not None and not preserved_matched_mtp.empty:
            matched_mtp_indices |= set(preserved_matched_mtp.index)
        matched_bank_indices = set(res.matched["_bank_index"].values) if not res.matched.empty else set()
        if preserved_matched_bank is not None and not preserved_matched_bank.empty:
            matched_bank_indices |= set(preserved_matched_bank.index)
        multi_mtp_indices = set(res.multi_matches.index) if not res.multi_matches.empty else set()

        # Build per-file bytes for individual downloads and for ZIP
        mtp_offsets = [0]
        for _name, mdf in mtp_files_data:
            mtp_offsets.append(mtp_offsets[-1] + len(mdf))
        mtp_file_bytes_list = []
        for i, (name, mdf) in enumerate(mtp_files_data):
            start = mtp_offsets[i]
            def _mtp_status(j):
                idx = start + j
                if idx in matched_mtp_indices:
                    return "Matched"
                if idx in multi_mtp_indices:
                    return "Multi-match"
                return "Unmatched"
            mtp_export = mdf.copy()
            mtp_export["Recon status"] = [ _mtp_status(j) for j in range(len(mdf)) ]
            control_col = mtp_control_choices[i]
            mtp_highlighted_bytes = _excel_with_control_column_highlight(
                mtp_export, "Recon status", control_col
            )
            mtp_file_bytes_list.append((name, mtp_highlighted_bytes))
            safe_name = _sanitize_sheet_name(name).rstrip("_") or "mtp"
            st.download_button(
                f"Download **{name}** (Recon status + highlighting)",
                data=mtp_highlighted_bytes,
                file_name=f"{safe_name}_with_results.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"dl_mtp_hl_{i}",
            )

        bank_offsets = [0]
        for _name, bdf in bank_files_data:
            bank_offsets.append(bank_offsets[-1] + len(bdf))
        bank_file_bytes_list = []
        for i, (name, bdf) in enumerate(bank_files_data):
            start = bank_offsets[i]
            statuses = [
                "Matched" if (start + j) in matched_bank_indices else "Unmatched"
                for j in range(len(bdf))
            ]
            bank_export = bdf.copy()
            bank_export["Recon status"] = statuses
            bank_highlighted_bytes = _excel_with_highlighted_rows(bank_export)
            bank_file_bytes_list.append((name, bank_highlighted_bytes))
            safe_name = _sanitize_sheet_name(name).rstrip("_") or "statement"
            st.download_button(
                f"Download **{name}** (Recon status + highlighting)",
                data=bank_highlighted_bytes,
                file_name=f"{safe_name}_with_results.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"dl_bank_hl_{i}",
            )

        # ZIP with results workbook + all uploaded files with results
        if mtp_file_bytes_list or bank_file_bytes_list:
            zip_buf = io.BytesIO()
            with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
                zf.writestr("mtp_reconciliation_results.xlsx", results_excel_bytes)
                for name, data in mtp_file_bytes_list:
                    safe = _sanitize_sheet_name(name).rstrip("_") or "mtp"
                    zf.writestr(f"uploaded_with_results/{safe}_with_results.xlsx", data)
                for name, data in bank_file_bytes_list:
                    safe = _sanitize_sheet_name(name).rstrip("_") or "statement"
                    zf.writestr(f"uploaded_with_results/{safe}_with_results.xlsx", data)
            zip_buf.seek(0)
            st.download_button(
                "Download all (results + each uploaded file with results as ZIP)",
                data=zip_buf.getvalue(),
                file_name="mtp_recon_all_results.zip",
                mime="application/zip",
                key="dl_all_zip",
            )

else:
    st.info("Upload at least one MTP file and one bank statement Excel file to continue.")
    if mtp_files_data:
        st.subheader("MTP file(s) preview")
        for name, df in mtp_files_data[:3]:
            st.caption(name)
            st.dataframe(_safe_for_display(df.head(5)), width="stretch")
    if bank_files_data:
        st.subheader("Bank statement(s) preview")
        for name, df in bank_files_data[:3]:
            st.caption(name)
            st.dataframe(_safe_for_display(df.head(5)), width="stretch")

# Developer acknowledgment (sidebar)
st.sidebar.markdown("---")
st.sidebar.caption(
    "Built by [**Erick D. Makilagi**](https://github.com/makilagied) · "
    "[@makilagied](https://github.com/makilagied)"
)
